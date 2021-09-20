import openpyxl
import xlrd

filename = r'./data/news0.xlsx'
book = xlrd.open_workbook(filename, encoding_override='utf-8')  # 读文件
# sheetnames = book.sheet_by_index(0)  # 获取读文件中所有的sheet，通过名字的方式

data = []
label = []
for i in range(0, 14):
    ws = book.sheet_by_index(i)  # 获取第一个sheet内容
    # 获取sheet的最大行数和列数
    rows = 200 ###每个类取1000条数据
    for r in range(1, ws.nrows):
        if r%700==0:
            row_data = ws.row_values(r)
            data.append(row_data[0].strip('\n'))
            label.append(int(row_data[1]))


####写入文件
outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
for row in range(1,len(data)+1):
    outws.cell(row, 1).value = data[row-1]  # 写文件
    outws.cell(row, 2).value = label[row-1]
saveExcel = "./data/datatest.xlsx"
outwb.save(saveExcel)  # 一定要记得保存