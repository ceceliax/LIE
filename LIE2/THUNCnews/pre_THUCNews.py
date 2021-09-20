import os
import os
import os.path
# rootdir="F:\研一下\学长论文\hjx\最终备份\O2M\my_xxx\THUNews\THUCNews"                                # 指明被遍历的文件夹
rootdir = "F:\研一下\学长论文\hjx\最终备份\O2M\my_xxx\\news\娱乐"
class_num = 14
# class_dic = {"科技": "0", "彩票": "1", "房产": "2", "股票": "3", "家居": "4", "教育": "5", "财经": "6",
#              "社会": "7", "时尚": "8", "时政": "9", "体育": "10", "星座": "11", "游戏": "12", "娱乐": "13"}
# sendict = {str(item): [] for item in range(class_num)}
# sendict = {"0": [], "1": [], "2": [], "3": [], "4": [], "5": [], "6": [], "7": [], "8": [], "9": [], "10": [],
    #            "11": [], "12": [], "13": [], "14": []}
caijing = []
for parent,dirnames,filenames in os.walk(rootdir):    #三个参数：分别返回1.父目录 2.所有文件夹名字（不含路径） 3.所有文件名字
    for filename in filenames:
        with open(os.path.join(parent,filename),encoding='utf-8') as f:
            line = f.readline()
            caijing.append(line)

# print(caijing)


import openpyxl
outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
for row in range(1,len(caijing)+1):
    outws.cell(row, 1).value = caijing[row-1]  # 写文件
    outws.cell(row, 2).value = 13
saveExcel = "news13.xlsx"
outwb.save(saveExcel)  # 一定要记得保存


# def readExel():
#     filename = r'D:\test.xlsx'
#     inwb = openpyxl.load_workbook(filename)  # 读文件
#     sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
#     ws = inwb.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容
#
#     # 获取sheet的最大行数和列数
#     rows = ws.max_row
#     cols = ws.max_column
#     for r in range(1,rows):
#         for c in range(1,cols):
#             print(ws.cell(r,c).value)
#         if r==10:
#             break






