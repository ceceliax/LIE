import numpy as np
import _pickle
import re
import torch
from torch.utils.data import DataLoader, TensorDataset

def get_data(path):
    data = []
    with open(path, encoding='utf-8') as f:
        while True:
            lines = f.readline().strip('\n')
            if not lines:
                break
                pass
            data.append(lines)
    return data

row_data = r'./atis/dev/seq.in'
row_label = r'./atis/dev/label'
row_data = get_data(row_data) #数据读入
row_label = get_data(row_label)
# label = set()
# for i in row_label:
#     label.add(i)



# label = set()
# for i in range(len(row_label)):
#     label.add(row_label[i])
# print(label)
# # for i in range(len(row_data)):
# #     w = open('./atis/dev.txt', 'a', encoding='utf-8')
# #     w.write(row_data[i]+'\t' + row_label[i] + '\n')
# #     w.close()
#
# import xlrd
# import jieba
#
#
#
# #### 生成停用词列表
#
classdict = {'UNK':'0',
             'atis_abbreviation':'1',
             'atis_aircraft':'2',
             'atis_aircraft#atis_flight#atis_flight_no':'3',
             'atis_airfare':'4',
             'atis_airline':'5',
             'atis_airline#atis_flight_no':'6',
             'atis_airport':'7',
             'atis_capacity':'8',
             'atis_cheapest':'9',
             'atis_city':'10',
             'atis_distance':'11',
             'atis_flight':'12',
             'atis_flight#atis_airfare':'13',
             'atis_flight_no':'14',
             'atis_flight_time':'15',
             'atis_ground_fare':'16',
             'atis_ground_service':'17',
             'atis_ground_service#atis_ground_fare':'18',
             'atis_meal':'19',
             'atis_quantity':'20',
             'atis_restriction':'21'
             }

# for i in range(len(row_data)):
#     w = open('./atis/dev.txt', 'a', encoding='utf-8')
#     try:
#         w.write(row_data[i]+'\t' + classdict[row_label[i]] + '\n')
#     except:
#         w.write(row_data[i] + '\t' + str(0) + '\n')
#     w.close()


#
# classdict={
# 'AddToPlaylist':'0',
# 'BookRestaurant':'1',
# 'GetWeather':'2',
# 'PlayMusic':'3',
# 'RateBook':'4',
# 'SearchCreativeWork':'5',
# 'SearchScreeningEvent':'6'
# }
#
#


import openpyxl
outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
for row in range(1,len(row_data)+1):
    outws.cell(row, 1).value = row_data[row-1]  # 写文件
    try:
        outws.cell(row, 2).value = int(classdict[row_label[row-1]])
    except:
        outws.cell(row, 2).value = 0
saveExcel = "./atis/atis_dev.xlsx"
# outwb.save(saveExcel)  # 一定要记得保存




label = [
        "what is mean,what does mean,code,explain,ff,qx,dfw,y,ap,f,h",
        "aircraft,flight,fly,type of, plane,kind of ,airplane, use",
        "aircraft,flight,flight number",
        "airfare,fares,the cost of a flight,fare",
        "airlines,which airlines,fly",
        "airline,flight number",
        "airpot,airpots",
        "how many,list the number of people,total,capacity,seating capacities, how many seats,carried on",
        "cheapest",
        "cities,city,denver,nationair,canadian,washington",
        "how long,distance,how far",
        "flights,from to,leave,fly,flight",
        "flight,fares",
        "the number of flights,flight numbers",
        "the schedule,time",
        "the rental car rates,the cost of limousine,ground transportation,how much,price,taxi,car,fare",
        "is there,transportation,types of ground transportation,kinds of,show me, rental cars,list,taxi,",
        "ground transportation,the cost of,fare"
        "a meal,meals",
        "how many flights,how many cities"
        "restrictions,restriction"

    ]